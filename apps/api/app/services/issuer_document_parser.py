from __future__ import annotations

import io
import math
import re
import unicodedata
from collections import defaultdict
from datetime import UTC, date, datetime
from html.parser import HTMLParser
from typing import Any, Iterable, Literal

from openpyxl import load_workbook
from pypdf import PdfReader

from app.schemas.fundamentals import (
    FinancialPeriod,
    FinancialSource,
    IssuerDocumentCandidate,
)


StatementType = Literal["income", "cashflow", "balance", "unknown"]


MONTHS: dict[str, int] = {
    "january": 1,
    "janvier": 1,
    "february": 2,
    "feb": 2,
    "fevrier": 2,
    "février": 2,
    "march": 3,
    "mars": 3,
    "april": 4,
    "avril": 4,
    "may": 5,
    "mai": 5,
    "june": 6,
    "juin": 6,
    "july": 7,
    "juillet": 7,
    "august": 8,
    "aout": 8,
    "août": 8,
    "september": 9,
    "septembre": 9,
    "october": 10,
    "octobre": 10,
    "november": 11,
    "novembre": 11,
    "december": 12,
    "decembre": 12,
    "décembre": 12,
}

MONTH_PATTERN = "|".join(
    sorted(
        (re.escape(month) for month in MONTHS),
        key=len,
        reverse=True,
    )
)

FULL_DATE_RE = re.compile(
    rf"\b({MONTH_PATTERN})\s+(\d{{1,2}})(?:st|nd|rd|th)?[,]?\s+"
    r"(20\d{2}|19\d{2})\b",
    re.IGNORECASE,
)
CONTEXT_DATE_RE = re.compile(
    rf"\b({MONTH_PATTERN})\s+(\d{{1,2}})(?:st|nd|rd|th)?\b",
    re.IGNORECASE,
)
YEAR_RE = re.compile(r"\b(20\d{2}|19\d{2})\b")
NUMBER_RE = re.compile(
    r"(?<![A-Za-z])"
    r"("
    r"\(?\s*[-+]?\$?\s*"
    r"(?:\d{1,3}(?:,\d{3})+|\d+)"
    r"(?:\.\d+)?\s*\)?"
    r"|-"
    r")"
    r"(?![A-Za-z])"
)
ISO_DATE_RE = re.compile(
    r"\b(20\d{2}|19\d{2})-(\d{2})-(\d{2})\b"
)

STATEMENT_HEADINGS: list[tuple[StatementType, tuple[str, ...]]] = [
    (
        "cashflow",
        (
            "statement of cash flows",
            "statements of cash flows",
            "consolidated cash flows",
            "etat des flux de tresorerie",
            "etats des flux de tresorerie",
            "flux de tresorerie consolides",
        ),
    ),
    (
        "balance",
        (
            "statement of financial position",
            "statements of financial position",
            "balance sheet",
            "balance sheets",
            "etat de la situation financiere",
            "etats de la situation financiere",
            "bilan consolide",
            "bilans consolides",
        ),
    ),
    (
        "income",
        (
            "statement of earnings",
            "statements of earnings",
            "statement of income",
            "statements of income",
            "statement of operations",
            "statements of operations",
            "consolidated results of operations",
            "etat des resultats",
            "etats des resultats",
            "etat du resultat",
            "comptes de resultat",
        ),
    ),
]

# Longest and most specific aliases are listed first.
ALIASES: dict[StatementType, dict[str, tuple[str, ...]]] = {
    "income": {
        "total_revenue": (
            "operating revenues",
            "total revenues",
            "revenue",
            "revenues",
            "sales",
            "chiffre d affaires",
            "produits d exploitation",
        ),
        "cost_of_revenue": (
            "cost of revenues",
            "cost of revenue",
            "cost of sales",
            "cout des ventes",
            "cout des produits",
        ),
        "gross_profit": (
            "gross profit",
            "gross margin",
            "benefice brut",
            "marge brute",
        ),
        "research_development": (
            "research and development expenses",
            "research and development expense",
            "research development",
            "frais de recherche et developpement",
        ),
        "selling_general_administrative": (
            "selling general and administrative expenses",
            "selling general administrative expenses",
            "general and administrative expenses",
            "frais de vente generaux et administratifs",
        ),
        "total_operating_expenses": (
            "total operating expenses",
            "operating expenses",
            "charges d exploitation",
        ),
        "operating_income": (
            "operating earnings",
            "operating income",
            "income from operations",
            "resultat d exploitation",
            "benefice d exploitation",
        ),
        "interest_expense": (
            "interest on long term debt",
            "interest expense",
            "finance costs",
            "charges d interets",
            "frais financiers",
        ),
        "income_before_tax": (
            "earnings before income taxes",
            "income before income taxes",
            "income before taxes",
            "profit before tax",
            "resultat avant impots",
            "benefice avant impots",
        ),
        "income_tax_expense": (
            "income tax expense",
            "provision for income taxes",
            "charge d impots",
            "impots sur le resultat",
        ),
        "net_income": (
            "net earnings attributable to common shareholders",
            "net income attributable to shareholders",
            "net earnings",
            "net income",
            "profit for the period",
            "resultat net",
            "benefice net",
        ),
        "basic_eps": (
            "basic earnings per share",
            "basic net income per share",
            "resultat de base par action",
            "benefice de base par action",
        ),
        "diluted_eps": (
            "diluted earnings per share",
            "diluted net income per share",
            "resultat dilue par action",
            "benefice dilue par action",
        ),
        "diluted_average_shares": (
            "weighted average diluted shares",
            "weighted average number of diluted shares",
            "nombre moyen pondere dilue d actions",
        ),
    },
    "cashflow": {
        "operating_cash_flow": (
            "net cash provided by operating activities",
            "cash provided by operating activities",
            "cash from operating activities",
            "cash used in operating activities",
            "cash used in from operating activities",
            "flux de tresorerie lies aux activites d exploitation",
            "tresorerie provenant des activites d exploitation",
        ),
        "capital_expenditure": (
            "additions to property plant and equipment",
            "purchases of property plant and equipment",
            "purchase of property plant and equipment",
            "capital expenditures",
            "acquisition d immobilisations corporelles",
            "ajouts aux immobilisations corporelles",
            "depenses en immobilisations",
        ),
        "dividends_paid": (
            "dividends paid",
            "cash dividends paid",
            "dividendes verses",
        ),
        "share_repurchases": (
            "repurchase of common shares",
            "repurchase of shares",
            "share repurchases",
            "rachat d actions ordinaires",
            "rachats d actions",
        ),
        "depreciation_amortization": (
            "depreciation and amortization",
            "depreciation depletion and amortization",
            "amortissement des immobilisations corporelles et incorporelles",
        ),
    },
    "balance": {
        "total_cash": (
            "cash and cash equivalents",
            "cash cash equivalents and short term investments",
            "tresorerie et equivalents de tresorerie",
            "cash",
            "tresorerie",
        ),
        "current_assets": (
            "total current assets",
            "current assets",
            "total de l actif courant",
            "actifs courants",
        ),
        "total_assets": (
            "total assets",
            "total de l actif",
            "actif total",
        ),
        "current_liabilities": (
            "total current liabilities",
            "current liabilities",
            "total du passif courant",
            "passifs courants",
        ),
        "total_liabilities": (
            "total liabilities excluding equity",
            "total liabilities",
            "total du passif",
            "passif total",
        ),
        "stockholder_equity": (
            "total shareholders equity",
            "shareholders equity",
            "stockholders equity",
            "total equity",
            "equity",
            "capitaux propres",
        ),
        "_current_debt": (
            "current portion of long term debt",
            "current portion of debt",
            "short term debt",
            "portion courante de la dette a long terme",
            "dette a court terme",
        ),
        "_long_term_debt": (
            "long term debt excluding current portion",
            "long term debt",
            "non current debt",
            "dette a long terme",
        ),
        "total_debt": (
            "total debt",
            "dette totale",
        ),
    },
}

BLOCKED_PREFIXES = (
    "cash at beginning",
    "cash at end",
    "net increase in cash",
    "net decrease in cash",
    "effect of foreign exchange",
    "total liabilities and equity",
    "total liabilities and shareholders equity",
)


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.lines: list[str] = []
        self._buffer: list[str] = []
        self._break_tags = {
            "tr",
            "p",
            "div",
            "li",
            "br",
            "h1",
            "h2",
            "h3",
            "h4",
        }

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        if tag.lower() in self._break_tags:
            self._flush()

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in self._break_tags:
            self._flush()

    def handle_data(self, data: str) -> None:
        cleaned = " ".join(data.split())
        if cleaned:
            self._buffer.append(cleaned)

    def _flush(self) -> None:
        if self._buffer:
            self.lines.append(" ".join(self._buffer))
            self._buffer = []

    def close(self) -> None:
        self._flush()
        super().close()


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(
        character
        for character in value
        if not unicodedata.combining(character)
    )
    value = value.lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9$%().,+\- ]+", " ", value)
    return " ".join(value.split())


def parse_number(value: str) -> float | None:
    token = value.strip()
    if not token or token == "-":
        return None

    negative = token.startswith("(") and token.endswith(")")
    token = (
        token.replace("$", "")
        .replace(",", "")
        .replace(" ", "")
        .replace("(", "")
        .replace(")", "")
    )

    try:
        parsed = float(token)
    except ValueError:
        return None

    if negative:
        parsed = -abs(parsed)

    return parsed if math.isfinite(parsed) else None


def extract_numbers(value: str) -> list[float | None]:
    return [
        parse_number(match.group(1))
        for match in NUMBER_RE.finditer(value)
    ]


def _date_from_parts(
    month_name: str,
    day_value: str,
    year_value: str,
) -> datetime | None:
    month = MONTHS.get(month_name.lower())
    if month is None:
        return None
    try:
        return datetime(
            int(year_value),
            month,
            int(day_value),
            tzinfo=UTC,
        )
    except ValueError:
        return None


def detect_period_dates(lines: list[str]) -> list[datetime]:
    header = " ".join(lines[:35])
    output: list[datetime] = []

    for match in ISO_DATE_RE.finditer(header):
        try:
            output.append(
                datetime(
                    int(match.group(1)),
                    int(match.group(2)),
                    int(match.group(3)),
                    tzinfo=UTC,
                )
            )
        except ValueError:
            pass

    for match in FULL_DATE_RE.finditer(header):
        parsed = _date_from_parts(
            match.group(1),
            match.group(2),
            match.group(3),
        )
        if parsed is not None:
            output.append(parsed)

    if not output:
        context = CONTEXT_DATE_RE.search(header)
        years = [
            int(year)
            for year in YEAR_RE.findall(header)
        ]
        if context and years:
            for year in years[:3]:
                parsed = _date_from_parts(
                    context.group(1),
                    context.group(2),
                    str(year),
                )
                if parsed is not None:
                    output.append(parsed)

    unique: list[datetime] = []
    seen: set[date] = set()
    for parsed in output:
        if parsed.date() not in seen:
            seen.add(parsed.date())
            unique.append(parsed)

    return unique[:3]


def detect_currency_and_multiplier(
    lines: list[str],
) -> tuple[str | None, float]:
    sample = normalize_text(" ".join(lines[:45]))

    if (
        "united states dollars" in sample
        or "u.s. dollars" in sample
        or "dollars americains" in sample
        or "usd" in sample
    ):
        currency = "USD"
    elif (
        "canadian dollars" in sample
        or "dollars canadiens" in sample
        or "cad" in sample
    ):
        currency = "CAD"
    else:
        currency = None

    if (
        "in millions" in sample
        or "en millions" in sample
    ):
        multiplier = 1_000_000.0
    elif (
        "in thousands" in sample
        or "en milliers" in sample
    ):
        multiplier = 1_000.0
    else:
        multiplier = 1.0

    return currency, multiplier


def statement_type(line: str) -> StatementType | None:
    normalized = normalize_text(line)
    for kind, headings in STATEMENT_HEADINGS:
        if any(heading in normalized for heading in headings):
            return kind
    return None


def statement_sections(
    lines: list[str],
) -> list[tuple[StatementType, list[str]]]:
    starts: list[tuple[int, StatementType]] = []

    for index, line in enumerate(lines):
        kind = statement_type(line)
        if kind is not None:
            starts.append((index, kind))

    if not starts:
        return [("unknown", lines)]

    sections: list[tuple[StatementType, list[str]]] = []

    for position, (start, kind) in enumerate(starts):
        end = (
            starts[position + 1][0]
            if position + 1 < len(starts)
            else min(len(lines), start + 220)
        )
        sections.append((kind, lines[start:end]))

    return sections


def _alias_match(
    normalized_line: str,
    aliases: Iterable[str],
) -> tuple[str, int] | None:
    comparable = re.sub(
        r"[().,\-]+",
        " ",
        normalized_line,
    )
    comparable = " ".join(comparable.split())

    if any(
        comparable.startswith(prefix)
        for prefix in BLOCKED_PREFIXES
    ):
        return None

    best: tuple[str, int] | None = None

    for alias in aliases:
        index = comparable.find(alias)
        if index < 0 or index > 18:
            continue
        if best is None or len(alias) > len(best[0]):
            best = (alias, index)

    return best


def _field_values(
    line: str,
    alias: str,
    periods: int,
    multiplier: float,
    field: str,
) -> list[float | None]:
    del alias
    numbers = extract_numbers(line)

    if not numbers:
        return []

    selected = (
        numbers[-periods:]
        if periods > 0
        else numbers[-1:]
    )

    if field in {
        "basic_eps",
        "diluted_eps",
        "diluted_average_shares",
    }:
        scale = 1.0
    else:
        scale = multiplier

    output: list[float | None] = []
    for number in selected:
        output.append(
            number * scale
            if number is not None
            else None
        )
    return output


def _derive_period(
    values: dict[str, Any],
) -> dict[str, Any]:
    current_debt = values.pop("_current_debt", None)
    long_debt = values.pop("_long_term_debt", None)

    if values.get("total_debt") is None and (
        current_debt is not None
        or long_debt is not None
    ):
        values["total_debt"] = (
            (current_debt or 0)
            + (long_debt or 0)
        )

    if (
        values.get("net_debt") is None
        and values.get("total_debt") is not None
        and values.get("total_cash") is not None
    ):
        values["net_debt"] = (
            values["total_debt"]
            - values["total_cash"]
        )

    if (
        values.get("free_cash_flow") is None
        and values.get("operating_cash_flow") is not None
        and values.get("capital_expenditure") is not None
    ):
        values["free_cash_flow"] = (
            values["operating_cash_flow"]
            + values["capital_expenditure"]
        )

    if (
        values.get("ebitda") is None
        and values.get("operating_income") is not None
        and values.get("depreciation_amortization") is not None
    ):
        values["ebitda"] = (
            values["operating_income"]
            + abs(values["depreciation_amortization"])
        )

    revenue = values.get("total_revenue")

    def ratio(field: str) -> float | None:
        value = values.get(field)
        if value is None or revenue in (None, 0):
            return None
        return value / revenue * 100

    values["gross_margin"] = ratio("gross_profit")
    values["operating_margin"] = ratio(
        "operating_income"
    )
    values["net_margin"] = ratio("net_income")
    values["free_cash_flow_margin"] = ratio(
        "free_cash_flow"
    )

    return values


class FinancialDocumentParser:
    max_pdf_pages = 45
    max_xlsx_rows_per_sheet = 650
    max_xlsx_columns = 36

    def parse_text(
        self,
        text: str,
        document: IssuerDocumentCandidate,
    ) -> list[FinancialPeriod]:
        lines = [
            " ".join(line.split())
            for line in text.splitlines()
            if " ".join(line.split())
        ]

        if not lines:
            return []

        document_kind: Literal[
            "annual",
            "quarterly",
        ] = (
            "annual"
            if document.document_type == "annual"
            else "quarterly"
        )

        merged: dict[
            tuple[date, str],
            dict[str, Any],
        ] = {}

        for kind, section in statement_sections(lines):
            if kind == "unknown":
                continue

            dates = detect_period_dates(section)
            if not dates:
                dates = detect_period_dates(lines[:80])
            if not dates:
                continue

            currency, multiplier = (
                detect_currency_and_multiplier(
                    section
                )
            )
            if currency is None:
                currency, global_multiplier = (
                    detect_currency_and_multiplier(
                        lines[:100]
                    )
                )
                if multiplier == 1:
                    multiplier = global_multiplier

            aliases = ALIASES[kind]

            for line in section:
                normalized = normalize_text(line)

                matches: list[
                    tuple[int, str, str]
                ] = []

                for field, field_aliases in aliases.items():
                    matched = _alias_match(
                        normalized,
                        field_aliases,
                    )
                    if matched is not None:
                        alias, _ = matched
                        matches.append(
                            (
                                len(alias),
                                field,
                                alias,
                            )
                        )

                if not matches:
                    continue

                _, field, alias = max(
                    matches,
                    key=lambda item: item[0],
                )
                values = _field_values(
                    line,
                    alias,
                    len(dates),
                    multiplier,
                    field,
                )
                if not values:
                    continue

                # Keep the most recent N values aligned to the
                # most recent dates shown in the statement header.
                aligned_dates = dates[: len(values)]

                for parsed_date, value in zip(
                    aligned_dates,
                    values,
                ):
                    if value is None:
                        continue

                    key = (
                        parsed_date.date(),
                        document_kind,
                    )
                    row = merged.setdefault(
                        key,
                        {
                            "period_end": parsed_date,
                            "period_type": document_kind,
                            "currency": currency,
                        },
                    )

                    # The first statement occurrence wins unless
                    # a later row provides a currently missing field.
                    if row.get(field) is None:
                        if field in {
                            "capital_expenditure",
                            "dividends_paid",
                            "share_repurchases",
                        }:
                            value = -abs(value)
                        row[field] = value

        source = FinancialSource(
            source_type="issuer_official_document",
            source_name=document.title,
            source_url=document.url,
            filed_at=document.published_at,
            form=(
                "Issuer annual financial document"
                if document_kind == "annual"
                else "Issuer interim financial document"
            ),
            confidence="official",
        )

        output: list[FinancialPeriod] = []

        for _, row in sorted(
            merged.items(),
            key=lambda pair: pair[0][0],
            reverse=True,
        ):
            row = _derive_period(row)
            row["source"] = source

            populated = sum(
                row.get(field) is not None
                for field in (
                    "total_revenue",
                    "operating_income",
                    "net_income",
                    "operating_cash_flow",
                    "capital_expenditure",
                    "total_cash",
                    "total_debt",
                    "current_assets",
                    "current_liabilities",
                    "total_assets",
                    "total_liabilities",
                    "stockholder_equity",
                )
            )

            # Reject presentation-like documents that only happen to
            # contain one or two isolated financial values.
            if populated >= 4:
                output.append(FinancialPeriod(**row))

        return output

    def _pdf_text(self, content: bytes) -> str:
        reader = PdfReader(io.BytesIO(content))
        pages: list[str] = []

        for page in reader.pages[: self.max_pdf_pages]:
            try:
                text = page.extract_text(
                    extraction_mode="layout"
                )
            except TypeError:
                text = page.extract_text()
            except Exception:  # noqa: BLE001
                text = ""

            if text:
                pages.append(text)

        return "\n".join(pages)

    @staticmethod
    def _xlsx_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            return f"{value:.10g}"
        return str(value).strip()

    def _xlsx_text(self, content: bytes) -> str:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
        lines: list[str] = []

        for worksheet in workbook.worksheets:
            lines.append(worksheet.title)

            for index, row in enumerate(
                worksheet.iter_rows(
                    values_only=True,
                    max_col=self.max_xlsx_columns,
                )
            ):
                if index >= self.max_xlsx_rows_per_sheet:
                    break

                values = [
                    self._xlsx_cell(value)
                    for value in row
                ]
                values = [value for value in values if value]

                if values:
                    lines.append(" | ".join(values))

        workbook.close()
        return "\n".join(lines)

    @staticmethod
    def _html_text(content: bytes) -> str:
        parser = _HTMLTextExtractor()
        parser.feed(
            content.decode(
                "utf-8",
                errors="replace",
            )
        )
        parser.close()
        return "\n".join(parser.lines)

    def parse_bytes(
        self,
        content: bytes,
        document: IssuerDocumentCandidate,
    ) -> list[FinancialPeriod]:
        if document.document_format == "pdf":
            text = self._pdf_text(content)
        elif document.document_format in {
            "xlsx",
            "xls",
        }:
            text = self._xlsx_text(content)
        elif document.document_format == "html":
            text = self._html_text(content)
        else:
            return []

        return self.parse_text(text, document)


financial_document_parser = FinancialDocumentParser()
